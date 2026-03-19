"""Constantes et paramètres globaux de la réconciliation."""

import re
from pathlib import Path

from openpyxl.styles import PatternFill

# ─── Chemins ──────────────────────────────────────────────────────────────────

DATA_DIR = Path("data")
OUTPUT_DIR = Path("output")

ERP_FILE = DATA_DIR / "Export_invoices.xlsx"

# Fichier banque : prend automatiquement le plus récent dans data/
_bank_candidates = sorted(
    DATA_DIR.glob("Export des ecritures*.xlsx"), reverse=True
)
BANK_FILE: Path = (
    _bank_candidates[0]
    if _bank_candidates
    else DATA_DIR / "Export des ecritures.xlsx"
)

CLIENT_MAPPING_PATH = DATA_DIR / "client_mapping.csv"

# ─── Patterns ─────────────────────────────────────────────────────────────────

# Capte NNNN-NNN, NNN-NNN (préfixe court), NNNN NNN (espace)
INVOICE_PATTERN = re.compile(r"(?<!\d)(\d{3,4})[-\s](\d{3,5})(?!\d)")

# ─── Mapping comptes bancaires ─────────────────────────────────────────────────

ACCOUNT_MAPPING: dict[str, str] = {
    "BPRI": "1 ELEVEN BPRI",
    "LCL": "3 eleven LCL",
    "CA": "2 ELEVEN CA",
}

# ─── Statuts ERP ──────────────────────────────────────────────────────────────

# Factures à rapprocher (non encore payées)
UNPAID_STATUSES: set[str] = {"Emise", "Envoyée", "Non émise"}

# Statuts exclus du rapprochement (factures annulées)
CANCELLED_STATUSES: set[str] = {"Annulée"}

# Type ERP indiquant une note de crédit (avoir) — champ "Type", pas "Statut"
AVOIR_TYPE: str = "Avoir"

# ─── Règles d'exclusion des lignes bancaires ──────────────────────────────────

# Types d'opération ignorés silencieusement (pas de label dans l'output)
EXCLUDED_OPERATION_TYPES: set[str] = {"12", "63", "77", "99"}

# Types d'opération exclus mais étiquetés dans la colonne Q
LABELED_OPERATION_TYPES: dict[str, str] = {
    "13": "TRESO",  # Virement de trésorerie reçu
}

# Mots-clés exclus silencieusement
SILENT_LIBELLE_KEYWORDS: list[str] = [
    "11 INVEST",
    "FACTOR",
]

# Mots-clés exclus mais étiquetés dans la colonne Q de l'output
LABELED_LIBELLE_KEYWORDS: dict[str, str] = {
    "CPAM": "CPAM",
    "URSSAF": "URSSAF",
    "SIE": "SIE",
    "OPERAT.COMPTE A PREAVIS": "OPERAT.COMPTE A PREAVIS",
    "TRESO": "TRESO",
    "VIR VAL COMPENS": "TRESO",
}

# ─── Options de réconciliation ────────────────────────────────────────────────

# Activer la déduction d'avoirs dans les combos de paiement.
# Formule : Σfactures − Σavoirs = crédit_banque
# Mettre à False pour ignorer les avoirs dans les combos.
ENABLE_AVOIRS_COMBO: bool = True

# ─── Tolerances ───────────────────────────────────────────────────────────────

# Tolérance en jours pour le rapprochement par date de paiement ERP (passe 0)
DATE_TOLERANCE_DAYS: int = 7

# Tolérance en jours pour la sous-passe B3 (montant + date paiement ERP)
DATE_TOLERANCE_AMOUNT_DATE: int = 7

# ─── Scoring (max = 100) ──────────────────────────────────────────────────────

SCORE_INVOICE_NUM: int = 40   # numéro de facture trouvé dans le libellé
SCORE_AMOUNT_EXACT: int = 30  # montant exact
SCORE_ACCOUNT: int = 15       # même compte bancaire
SCORE_CLIENT_MAX: int = 15    # similarité client >= 80

THRESHOLD_AUTO: int = 80    # score >= 80 → rapprochement automatique
THRESHOLD_REVIEW: int = 40  # score >= 40 → à vérifier manuellement

# ─── Couleurs Excel ───────────────────────────────────────────────────────────

# vert : rapproché automatiquement
FILL_AUTO = PatternFill("solid", fgColor="C6EFCE")
# jaune : à vérifier
FILL_REVIEW = PatternFill("solid", fgColor="FFEB9C")
# rouge : non rapproché
FILL_NONE = PatternFill("solid", fgColor="FFC7CE")
