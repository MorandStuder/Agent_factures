"""Génération des fichiers Excel de sortie (ERP enrichi et relevé bancaire enrichi)."""

import copy
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

from .config import (
    FILL_AUTO,
    FILL_NONE,
    FILL_REVIEW,
    THRESHOLD_AUTO,
    UNPAID_STATUSES,
)
from .loaders import normalize_amount, parse_date, to_date

DATE_FORMAT = "DD/MM/YYYY"


def extend_worksheet_tables(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    new_col_names: list[str],
) -> None:
    """Étend les tableaux Excel (ListObjects) existants pour inclure de nouvelles colonnes.

    Met à jour la référence du tableau (ex: "A1:P200" → "A1:V200") et ajoute les
    définitions TableColumn correspondantes.

    Args:
        ws: feuille de calcul openpyxl.
        new_col_names: noms des nouvelles colonnes à ajouter en fin de tableau.
    """
    for table in ws.tables.values():
        m = re.match(r"(\$?[A-Za-z]+)(\$?\d+):(\$?[A-Za-z]+)(\$?\d+)", table.ref)
        if not m:
            continue
        start_col_str, start_row_str, end_col_str, end_row_str = m.groups()
        start_col = start_col_str.lstrip("$").upper()
        start_row = start_row_str.lstrip("$")
        end_col = end_col_str.lstrip("$").upper()
        end_row = end_row_str.lstrip("$")

        end_idx = column_index_from_string(end_col)
        # Utilise max_column de la feuille comme base : le tableau peut se terminer
        # avant la dernière colonne de données
        base_idx = max(end_idx, ws.max_column)
        new_end_col = get_column_letter(base_idx + len(new_col_names))
        table.ref = f"{start_col}{start_row}:{new_end_col}{end_row}"

        existing_cols: list[TableColumn] = list(table.tableColumns) if table.tableColumns else []
        next_id = (max(c.id for c in existing_cols) + 1) if existing_cols else (end_idx + 1)
        for i, name in enumerate(new_col_names):
            table.tableColumns.append(TableColumn(id=next_id + i, name=name))


def write_erp_output(
    source_path: Path,
    all_invoices: list[dict],
    erp_matches: dict[int, dict],
    output_path: Path,
) -> None:
    """Écrit le fichier ERP enrichi avec les colonnes de réconciliation.

    Ajoute 6 colonnes après les colonnes existantes :
    - Date paiement constatée
    - Référence bancaire
    - Score confiance
    - Note réconciliation
    - Libellé bancaire
    - Montant crédit

    Colorie chaque ligne selon le score (vert >= 80, jaune >= 40, rouge = non rapproché).
    Auto-dimensionne toutes les colonnes et démasque les colonnes cachées.

    Args:
        source_path: chemin du fichier ERP source.
        all_invoices: toutes les factures ERP.
        erp_matches: résultats de réconciliation par _row_idx.
        output_path: chemin du fichier de sortie.
    """
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active
    max_col = ws.max_column

    new_cols = [
        "Date paiement constatée",
        "Référence bancaire",
        "Score confiance",
        "Note réconciliation",
        "Libellé bancaire",
        "Montant crédit",
    ]

    # En-têtes des nouvelles colonnes
    for i, name in enumerate(new_cols):
        ws.cell(row=1, column=max_col + 1 + i, value=name)

    # Auto-dimensionnement + démasquage de toutes les colonnes
    for col_cells in ws.iter_cols():
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        dim = ws.column_dimensions[col_letter]
        dim.hidden = False
        dim.width = min(max(max_len + 2, 10), 50)

    unpaid_idx = {
        inv["_row_idx"]
        for inv in all_invoices
        if inv.get("Statut") in UNPAID_STATUSES
    }

    for inv in all_invoices:
        row_idx = inv["_row_idx"]
        match = erp_matches.get(row_idx)

        if match:
            score = match["score"]
            fill = FILL_AUTO if score >= THRESHOLD_AUTO else FILL_REVIEW
            values = [
                match["date_constatee"] or "",
                match["ref_bancaire"] or "",
                score,
                match["note"] or "",
                match.get("bank_libelle") or "",
                match.get("bank_credit") or "",
            ]
        elif row_idx in unpaid_idx:
            fill = FILL_NONE
            values = ["", "", 0, "Non rapproché", "", ""]
        else:
            continue  # facture payée sans correspondance : pas de coloriage

        DATE_IDX = 0        # "Date paiement constatée"
        MONTANT_CREDIT_IDX = 5  # "Montant crédit"
        for i, val in enumerate(values):
            # Convertit les dates string en datetime pour format natif Excel
            if i == DATE_IDX and val:
                val = to_date(val) or val
            cell = ws.cell(row=row_idx, column=max_col + 1 + i, value=val)
            cell.fill = fill
            if i == DATE_IDX and isinstance(cell.value, datetime):
                cell.number_format = DATE_FORMAT
            elif i == MONTANT_CREDIT_IDX and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"

    # ── Conversion des colonnes date (strings → datetime Excel) ─────────────
    # Identifie toutes les colonnes dont le header contient "Date"
    # (exclut "Période" qui est une plage texte libre)
    date_col_indices = [
        col for col in range(1, ws.max_column + 1)
        if "Date" in str(ws.cell(1, col).value or "")
    ]
    for col in date_col_indices:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                converted = to_date(cell.value)
                if converted:
                    cell.value = converted
                    cell.number_format = DATE_FORMAT

    # ── Tableau Excel vert clair (TableStyleLight4) ───────────────────────────
    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    erp_table = Table(
        displayName="FacturesERP",
        ref=table_ref,
    )
    erp_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(erp_table)

    # ── Masquer les colonnes B C D E (peu utiles en consultation) ────────────
    for col_letter in ("B", "C", "D", "E"):
        ws.column_dimensions[col_letter].hidden = True

    # ── Figer les volets en G2 ────────────────────────────────────────────────
    ws.freeze_panes = "G2"

    wb.save(output_path)


def write_bank_output(
    source_path: Path,
    bank_matches: dict[int, list[str]],
    bank_row_to_invoices: dict[int, list[dict]],
    labeled_exclusions: dict[int, str],
    bank_row_scores: dict[int, int],
    output_path: Path,
) -> None:
    """Écrit le relevé bancaire enrichi avec les colonnes de réconciliation.

    Étend le tableau Excel existant avec 7 nouvelles colonnes :
    - Factures rapprochées
    - Client(s)
    - Projet(s)
    - Montant(s) TTC
    - Statut(s)
    - Date(s) émission
    - Score confiance

    Les headers des nouvelles colonnes héritent du style de la colonne A.
    Les lignes rapprochées sont colorées en vert (>= 80) ou jaune (< 80).

    Args:
        source_path: chemin du fichier banque source.
        bank_matches: dict {_row_idx banque -> [numéros de factures]}.
        bank_row_to_invoices: dict {_row_idx banque -> [dicts factures]}.
        labeled_exclusions: dict {_row_idx -> label} pour les lignes exclues étiquetées.
        bank_row_scores: dict {_row_idx banque -> score de confiance}.
        output_path: chemin du fichier de sortie.
    """
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active
    max_col = ws.max_column

    new_cols = [
        "Factures rapprochées",
        "Client(s)",
        "Projet(s)",
        "Montant(s) TTC",
        "Statut(s)",
        "Date(s) émission",
        "Score confiance",
    ]
    extend_worksheet_tables(ws, new_cols)

    # Style des headers : copié depuis la première cellule d'en-tête existante
    src = ws.cell(row=1, column=1)
    for i, name in enumerate(new_cols):
        cell = ws.cell(row=1, column=max_col + 1 + i, value=name)
        if src.font:
            cell.font = copy.copy(src.font)
        if src.fill and src.fill.fill_type:
            cell.fill = copy.copy(src.fill)

    # Lignes rapprochées
    for row_idx, nums in bank_matches.items():
        invs = bank_row_to_invoices.get(row_idx, [])
        score = bank_row_scores.get(row_idx, 0)
        fill = FILL_AUTO if score >= THRESHOLD_AUTO else FILL_REVIEW
        clients = ", ".join(str(i.get("Client", "") or "") for i in invs)
        projets = ", ".join(str(i.get("Nom du projet", "") or "") for i in invs)
        # Montant(s) TTC : somme numérique (nombre dans Excel)
        ttc_amounts = [normalize_amount(i.get("Total à facturer (TTC)")) or 0.0 for i in invs]
        ttc_value: float | str = sum(ttc_amounts) if ttc_amounts else ""
        statuts = ", ".join(str(i.get("Statut", "") or "") for i in invs)
        # Date(s) émission : datetime si facture unique, sinon chaîne
        emission_dates = [i.get("Date de facturation réelle") for i in invs]
        if len(emission_dates) == 1:
            emissions: datetime | str = to_date(emission_dates[0]) or parse_date(emission_dates[0]) or ""
        else:
            emissions = ", ".join(parse_date(d) or "" for d in emission_dates)
        row_values = [", ".join(nums), clients, projets, ttc_value, statuts, emissions, score]
        MONTANT_TTC_IDX = 3   # "Montant(s) TTC"
        EMISSION_IDX = 5      # "Date(s) émission"
        for i, val in enumerate(row_values):
            cell = ws.cell(row=row_idx, column=max_col + 1 + i, value=val)
            cell.fill = fill
            if i == MONTANT_TTC_IDX and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
            elif i == EMISSION_IDX and isinstance(val, datetime):
                cell.number_format = DATE_FORMAT

    # Lignes exclues mais étiquetées (CPAM, URSSAF, SIE, TRESO, etc.)
    for row_idx, label in labeled_exclusions.items():
        if row_idx in bank_matches:
            continue  # déjà renseigné par le rapprochement
        ws.cell(row=row_idx, column=max_col + 1, value=label).fill = FILL_REVIEW

    # ── Formatage des colonnes source : dates C/D, montant F ─────────────────
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    date_src_cols = [
        i + 1 for i, h in enumerate(headers)
        if h and any(k in str(h) for k in ("Date comptable", "Date valeur"))
    ]
    credit_cols = [
        i + 1 for i, h in enumerate(headers)
        if h and "Credit" in str(h)
    ]
    for col in date_src_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                converted = to_date(cell.value)
                if converted:
                    cell.value = converted
                    cell.number_format = DATE_FORMAT
            elif isinstance(cell.value, datetime):
                cell.number_format = DATE_FORMAT
    for col in credit_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            val = normalize_amount(cell.value)
            if val is not None:
                cell.value = val
                cell.number_format = "#,##0.00"

    wb.save(output_path)
