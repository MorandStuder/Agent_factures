"""Chargement des fichiers sources et normalisation des données."""

import csv
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from .config import (
    ACCOUNT_MAPPING,
    CLIENT_MAPPING_PATH,
    DATE_TOLERANCE_DAYS,
    OUTPUT_DIR,
)


def normalize_amount(val: object) -> Optional[float]:
    """Convertit une valeur montant en float arrondi à 2 décimales.

    Args:
        val: valeur brute (int, float, str avec virgule ou point).

    Returns:
        Float arrondi à 2 décimales, ou None si la conversion échoue.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    s = str(val).strip().replace(",", ".").replace(" ", "").replace("\xa0", "")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def parse_date(val: object) -> Optional[str]:
    """Normalise une date en string JJ/MM/AAAA.

    Args:
        val: valeur brute (datetime ou string).

    Returns:
        String au format JJ/MM/AAAA, ou None.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    return str(val).strip()


def to_date(val: object) -> Optional[datetime]:
    """Convertit une valeur date en datetime pour comparaison.

    Tente les formats JJ/MM/AAAA, AAAA-MM-JJ et JJ/MM/AA.

    Args:
        val: valeur brute.

    Returns:
        datetime, ou None si la conversion échoue.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def dates_close(d1: Optional[str], d2: Optional[str]) -> bool:
    """Retourne True si les deux dates sont à au plus DATE_TOLERANCE_DAYS jours l'une de l'autre.

    Args:
        d1: première date (string ou None).
        d2: deuxième date (string ou None).

    Returns:
        True si les dates sont dans la tolérance configurée.
    """
    dt1, dt2 = to_date(d1), to_date(d2)
    if dt1 is None or dt2 is None:
        return False
    return abs((dt1 - dt2).days) <= DATE_TOLERANCE_DAYS


def payment_before_emission(
    bank_date_str: Optional[str],
    inv: dict,
) -> bool:
    """Retourne True si la date bancaire est antérieure à l'émission de la facture.

    Utilisé pour filtrer les faux positifs où un crédit précède la facture.

    Args:
        bank_date_str: date comptable du crédit bancaire.
        inv: ligne facture ERP.

    Returns:
        True si le crédit est antérieur à la date de facturation réelle.
    """
    bank_dt = to_date(bank_date_str)
    emission_dt = to_date(inv.get("Date de facturation réelle"))
    if bank_dt is None or emission_dt is None:
        return False
    return bank_dt < emission_dt


def resolve_bank_account(
    inv: dict,
    account_overrides: dict[str, str],
) -> str:
    """Retourne l'intitulé banque effectif pour une facture.

    Priorité : override client_mapping > champ Compte bancaire ERP.

    Args:
        inv: ligne facture ERP.
        account_overrides: dict erp_client -> code_compte (ex: "LCL").

    Returns:
        Intitulé du compte bancaire (ex: "3 eleven LCL"), ou "" si inconnu.
    """
    client = str(inv.get("Client", "") or "").strip()
    erp_acct = account_overrides.get(client) or str(
        inv.get("Compte bancaire", "") or ""
    ).strip()
    return ACCOUNT_MAPPING.get(erp_acct, "")


def load_client_mapping() -> tuple[dict[str, list[str]], dict[str, str]]:
    """Charge client_mapping.csv.

    Returns:
        name_map    : erp_client -> [bank_name, ...] (aliases banque, peut être multiple).
        account_map : erp_client -> code_compte (ex: "LCL") override.
    """
    if not CLIENT_MAPPING_PATH.exists():
        return {}, {}
    name_map: dict[str, list[str]] = {}
    account_map: dict[str, str] = {}
    with open(CLIENT_MAPPING_PATH, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            client = row["erp_client"]
            if row.get("bank_name"):
                name_map.setdefault(client, []).append(row["bank_name"])
            if row.get("bank_account"):
                account_map[client] = row["bank_account"]
    return name_map, account_map


def load_previous_comments() -> dict[tuple, str]:
    """Charge les commentaires manuels depuis le dernier fichier banque annoté.

    Cherche le fichier Banque_reconciliee_*.xlsx le plus récent dans output/.
    La clé est (date_comptable, libellé, crédit) pour une identification robuste.

    Returns:
        Dict {(date, libellé, crédit) -> commentaire}.
    """
    candidates = sorted(OUTPUT_DIR.glob("Banque_reconciliee_*.xlsx"), reverse=True)
    if not candidates:
        return {}
    latest = candidates[0]
    print(f"  Chargement commentaires depuis {latest.name}")
    wb = openpyxl.load_workbook(latest)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h else "" for h in rows[0]]

    try:
        idx_date = headers.index("Date comptable")
        idx_lib = headers.index("Libellé")
        idx_credit = headers.index("Credit")
        idx_comment = headers.index("Commentaire")
    except ValueError:
        return {}

    result: dict[tuple, str] = {}
    for row in rows[1:]:
        comment = row[idx_comment]
        if not comment:
            continue
        key = (
            str(row[idx_date] or ""),
            str(row[idx_lib] or ""),
            str(row[idx_credit] or ""),
        )
        result[key] = str(comment)
    print(f"    {len(result)} lignes avec commentaire chargées")
    return result


def load_previous_erp_matches() -> set[str]:
    """Charge les numéros de factures rapprochées lors du dernier run ERP.

    Cherche le fichier ERP_reconcilie_*.xlsx le plus récent dans output/ et
    retourne les numéros dont le Score confiance est > 0.

    Returns:
        Set des numéros de factures rapprochées au run précédent.
    """
    candidates = sorted(OUTPUT_DIR.glob("ERP_reconcilie_*.xlsx"), reverse=True)
    if not candidates:
        return set()
    latest = candidates[0]
    try:
        wb = openpyxl.load_workbook(latest, read_only=True, data_only=True)
        ws = wb.active
        headers = [
            str(c.value) if c.value else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        num_col = next(
            (i for i, h in enumerate(headers)
             if "Numéro" in h and "facture" in h.lower()),
            None,
        )
        score_col = next(
            (i for i, h in enumerate(headers) if "Score" in h),
            None,
        )
        if num_col is None or score_col is None:
            wb.close()
            return set()
        matched: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            score = row[score_col]
            num = str(row[num_col] or "").strip()
            if num and score and float(score) > 0:
                matched.add(num)
        wb.close()
        return matched
    except Exception:
        return set()


def load_workbook_data(path: Path) -> tuple[list[str], list[dict]]:
    """Charge un fichier xlsx, retourne (headers, lignes dict avec _row_idx).

    Args:
        path: chemin vers le fichier Excel.

    Returns:
        Tuple (liste des en-têtes, liste des lignes sous forme de dicts).
        Chaque dict contient une clé _row_idx = numéro de ligne Excel (base 2).
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h else "" for h in rows[0]]
    data = []
    for i, row in enumerate(rows[1:], start=2):
        d = {headers[j]: row[j] for j in range(len(headers))}
        d["_row_idx"] = i
        data.append(d)
    return headers, data
